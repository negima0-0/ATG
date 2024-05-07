import csv
import json
import configparser
import os
from datetime import datetime
from netmiko import ConnectHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import WorkbookAlreadySaved
from openpyxl.utils import get_column_letter


# config.iniからユーザ名とパスワードを取得する
def get_credentials():
    config = configparser.ConfigParser()
    config.read('config.ini')

    username = config.get('credentials', 'username')
    password = config.get('credentials', 'password')

    return username, password


# ユーザ名とパスワードを取得
username, password = get_credentials()


# セッションログ用のフォルダが存在しない場合作成する
if not os.path.exists('session_logs'):
    os.makedirs('session_logs')


# CSVファイルからホスト名とIPアドレスの対応を読み込む
hosts = []
with open('hosts.csv', 'r') as csvfile:
    reader = csv.DictReader(csvfile)
    for row in reader:
        if 'hostname' in row and 'ip_address' in row:
            hosts.append((row['hostname'], row['ip_address']))
        elif 'hostname' in row:
            hosts.append((row['hostname'], None))
        elif 'ip_address' in row:
            hosts.append((None, row['ip_address']))


# ホストごとにデータをエクセルに書き込む
def write_to_excel(host, interface_info):
    print(f"{host} の情報をエクセルに書き込んでいます。")
    try:
        wb = load_workbook('output.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    try:
        ws = wb[host]
    except KeyError:
        ws = wb.create_sheet(title=host)
        headers = ["Timestamp", "Interface", "Input Multicasts", "Output Multicasts", "Input Broadcasts", "Output Broadcasts", "Input Errors", "Input Drops", "Framing Drops", "Input Runts", "Input Discards", "Input L3 Incompletes", "Input L2 Channel Errors", "Input L2 Mismatch Timeouts", "Input FIFO Errors", "Input Resource Errors"]
        ws.append(headers)

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for interface_data in interface_info:
        ws.append([timestamp] + list(interface_data))

    try:
        wb.save('output.xlsx')
        print(f"ホスト: {hostname or ip_address} のデータをエクセルに書き込みました。")
    except PermissionError as e:
        print(f"エラー: {e} (output.xlsxが開かれているため追記できません)")
        return
    except:
        print("エクセルファイルが開かれているか、その他要因によって追記できません。")
        return


# JSONデータからインターフェースごとの情報を取り出す
def extract_interface_info(data):
    interface_info = []
    for interface in data['interface-information'][0]['physical-interface']:
        interface_name = interface.get('name', [{}])[0].get('data', 'No Data')

        # 対象のインターフェースに絞る
        if interface_name.startswith(('ae', 'ge', 'xe', 'irb', 'reth')):
            tail_drop_packets = interface.get('ingress-queue-counters', [{}])[0].get('input-multicasts', [{}])[0].get('data', 'No Data')
            input_multicasts = interface.get('ethernet-mac-statistics', [{}])[0].get('input-multicasts', [{}])[0].get('data', 'No Data')
            output_multicasts = interface.get('ethernet-mac-statistics', [{}])[0].get('output-multicasts', [{}])[0].get('data', 'No Data')
            input_broadcasts = interface.get('ethernet-mac-statistics', [{}])[0].get('input-broadcasts', [{}])[0].get('data', 'No Data')
            output_broadcasts = interface.get('ethernet-mac-statistics', [{}])[0].get('output-broadcasts', [{}])[0].get('data', 'No Data')
            input_errors = interface.get('input-error-list', [{}])[0].get('input-errors', [{}])[0].get('data', 'No Data')
            input_drops = interface.get('input-error-list', [{}])[0].get('input-drops', [{}])[0].get('data', 'No Data')
            framing_drops = interface.get('input-error-list', [{}])[0].get('framing-errors', [{}])[0].get('data', 'No Data')
            input_runts = interface.get('input-error-list', [{}])[0].get('input-runts', [{}])[0].get('data', 'No Data')
            input_discards = interface.get('input-error-list', [{}])[0].get('input-discards', [{}])[0].get('data', 'No Data')
            input_l3_incompletes = interface.get('input-error-list', [{}])[0].get('input-l3-incompletes', [{}])[0].get('data', 'No Data')
            input_l2_channel_errors = interface.get('input-error-list', [{}])[0].get('input-l2-channel-errors', [{}])[0].get('data', 'No Data')
            input_l2_mismatch_timeouts = interface.get('input-error-list', [{}])[0].get('input-l2-mismatch-timeouts', [{}])[0].get('data', 'No Data')
            input_fifo_errors = interface.get('input-error-list', [{}])[0].get('input-fifo-errors', [{}])[0].get('data', 'No Data')
            input_resource_errors = interface.get('input-error-list', [{}])[0].get('input-resource-errors', [{}])[0].get('data', 'No Data')
            interface_info.append((interface_name, input_multicasts, output_multicasts, input_broadcasts, output_broadcasts, input_errors, input_drops, framing_drops, input_runts, input_discards, input_l3_incompletes, input_l2_channel_errors, input_l2_mismatch_timeouts, input_fifo_errors, input_resource_errors))
    return interface_info


# ホストへの接続
def main():
    for hostname, ip_address in hosts:
        try:
            # ホストへのSSH接続（ホスト名での接続）
            with ConnectHandler(
                device_type='juniper_junos',
                ip=ip_address,
                hostname=hostname,
                username=username,
                password=password,
                port=22,  # SSHポート（デフォルトは22）
                global_delay_factor=2,  # コマンド実行後のディレイ係数
                timeout=30,  # レスポンスまでのタイムアウトを30秒に設定(デフォルトは5秒)
                session_log=f"session_logs/{hostname or ip_address}.log"
            ) as ssh:
                # コマンドを実行してJSONデータを取得
                json_output = ssh.send_command('show interfaces queue forwarding-class AF |no-more |display json')

        except Exception as e:
            print(e)
            # ホスト名での接続が失敗した場合、IPアドレスで再接続を試みる
            print(f"ホスト名での接続に失敗しました。IPアドレス {ip_address} を使用して再接続を試みます...")
            try:
                # ホストへのSSH接続（IPアドレスでの接続）
                with ConnectHandler(
                    device_type='juniper_junos',
                    ip=ip_address,
                    username=username,
                    password=password,
                    port=22,  # SSHポート（デフォルトは22）
                    global_delay_factor=2,  # コマンド実行後のディレイ係数
                    timeout = 30,  # レスポンスまでのタイムアウトを30秒に設定(デフォルトは5秒)
                    session_log=f"session_logs/{hostname or ip_address}.log"
                ) as ssh:
                    # コマンドを実行してJSONデータを取得
                    json_output = ssh.send_command('show interfaces queue forwarding-class AF |no-more |display json', read_timeout=90) #タイムアウト90秒

            except Exception as e:
                print(f"ホスト {hostname or ip_address} への接続中にエラーが発生しました:", e)
                continue

        # JSONデータからインターフェースごとの情報を取得
        interface_info = extract_interface_info(json.loads(json_output))

        # エクセルにデータを書き込む
        #  create_excel_file_if_not_exists()
        write_to_excel(hostname or ip_address, interface_info)

        # 結果の出力
        # print(f"ホスト: {hostname or ip_address} のデータをエクセルに書き込みました。")

'''
    # 結果の出力
    print(f"ホスト: {hostname or ip_address}")
    for interface in interface_info:
        interface_name, input_multicasts, output_multicasts, input_broadcasts, output_broadcasts, input_errors, input_drops, framing_drops, input_runts, input_discards, input_l3_incompletes, input_l2_channel_errors, input_l2_mismatch_timeouts, input_fifo_errors, input_resource_errors = interface
        print(f"インターフェース名: {interface_name}")
        print("マルチキャストパケット（入力）:", input_multicasts)
        print("マルチキャストパケット（出力）:", output_multicasts)
        print("ブロードキャストパケット（入力）:", input_broadcasts)
        print("ブロードキャストパケット（出力）:", output_broadcasts)
        print("input errors: ", input_errors)
        print("input drops: ", input_drops)
        print("framing drops: ", framing_drops)
        print("input runts: ", input_runts)
        print("input discards: ", input_discards)
        print("input L3 incompletes: ", input_l3_incompletes)
        print("input L2 channel errors: ", input_l2_channel_errors)
        print("input L2 mismatch timeouts: ", input_l2_mismatch_timeouts)
        print("input fifo errors: ", input_fifo_errors)
        print("input resource errors: ", input_resource_errors)
        print('-' * 50)
'''

if __name__ == "__main__":
    main()
