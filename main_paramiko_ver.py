import os
import sys
import csv
import json
import logging
import paramiko
import configparser
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# ロギングの設定
# エラーログ用のロガーを設定
error_logger = logging.getLogger('errorLogger')
error_logger.setLevel(logging.ERROR)
error_handler = logging.FileHandler('error.log', mode='a')
error_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
error_logger.addHandler(error_handler)

# SSHセッションログ用のロガーを設定
ssh_logger = logging.getLogger('sshLogger')
ssh_logger.setLevel(logging.DEBUG)
ssh_handler = logging.FileHandler('ssh_sessions.log', mode='a')
ssh_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
ssh_logger.addHandler(ssh_handler)
paramiko.util.log_to_file('paramiko.log', level='DEBUG')

# 設定ファイルから認証情報を取得する関数
def get_credentials():
    config_file = 'config.ini'
    if not os.path.exists(config_file):
        error_logger.error(f"エラー: 設定ファイル '{config_file}' が見つかりません。")
        sys.exit(1)
    config = configparser.ConfigParser()
    config.read(config_file)
    username = config.get('credentials', 'username')
    password = config.get('credentials', 'password')
    return username, password

# ホスト情報の読み込み
def load_hosts():
    hosts_file = 'hosts.csv'
    if not os.path.exists(hosts_file):
        error_logger.error(f"エラー: ホスト情報ファイル '{hosts_file}' が見つかりません。")
        sys.exit(1)
    hosts = []
    with open(hosts_file, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            if 'hostname' in row and 'ip_address' in row:
                hosts.append((row['hostname'], row['ip_address']))
    return hosts

# ssh_via_jump_host関数の修正版
def ssh_via_jump_host(jump_host_ip, jump_host_user, jump_host_pass, target_host_ip, username, password):
    try:
        print(f"踏み台ホスト {jump_host_ip} への接続を試みます...")
        jump_ssh = paramiko.SSHClient()
        jump_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        jump_ssh.connect(jump_host_ip, username=jump_host_user, password=jump_host_pass, port=22)
        print("踏み台ホストへの接続に成功しました。")
    except paramiko.AuthenticationException:
        error_logger.error("踏み台サーバへの認証失敗。")
        sys.exit(1)
    except paramiko.SSHException as e:
        error_logger.error(f"踏み台サーバへの接続エラー: {e}")
        sys.exit(1)

    try:
        print(f"{target_host_ip} に接続を試みます...")
        target_ssh = paramiko.SSHClient()
        target_ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target_ssh.connect(target_host_ip, port=22, username=username, password=password)
        print(f"{target_host_ip} への接続に成功しました。")
    except paramiko.AuthenticationException:
        error_logger.error("Junosホストへの認証失敗。")
        sys.exit(1)
    except paramiko.SSHException as e:
        error_logger.error(f"Junosホストへの接続エラー: {e}")
        sys.exit(1)
    # コマンド実行
    try:
        stdin, stdout, stderr = target_ssh.exec_command('show interfaces extensive | display json | no-more')
        json_output = stdout.read().decode('utf-8')  # 出力を文字列としてデコード

        # リソースのクリーンアップ
        target_ssh.close()
        jump_ssh.close()

        return json_output
    except Exception as e:
        logging.error(f"接続エラー: {e}")  # エラーメッセージをログに記録
        print(f"接続エラー: {e}")  # コンソールにも出力
        sys.exit(1)


# JSONデータからインターフェース情報を抽出
def extract_interface_info(data):
    interface_info = []
    for interface in data['interface-information'][0]['physical-interface']:
        # インターフェース名の抽出
        interface_name = interface.get('name', [{}])[0].get('data', 'No Data')
        if not interface_name.startswith(('ae', 'ge', 'xe', 'irb', 'reth')):
            continue
        # 必要なデータの抽出（存在しない場合はNoneを設定）
        input_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('input-packets', [{}])[0].get('data', 'None')
        output_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('output-packets', [{}])[0].get('data', 'None')
        input_unicasts = interface.get('ethernet-mac-statistics', [{}])[0].get('input-unicasts', [{}])[0].get('data', 'None')
        output_unicasts = interface.get('ethernet-mac-statistics', [{}])[0].get('output-unicasts', [{}])[0].get('data', 'None')
        input_broadcast_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('input-broadcasts', [{}])[0].get('data', 'None')
        output_broadcast_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('output-broadcasts', [{}])[0].get('data', 'None')
        input_multicast_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('input-broadcasts', [{}])[0].get('data', 'None')
        output_multicast_packets = interface.get('ethernet-mac-statistics', [{}])[0].get('output-broadcasts', [{}])[0].get('data', 'None')
        interface_info.append((interface_name, input_packets, output_packets, input_unicasts, output_unicasts, input_broadcast_packets, output_broadcast_packets, input_multicast_packets, output_multicast_packets))
    return interface_info

# エクセルにデータを書き込む関数
def write_to_excel(hostname, interface_info):
    excel_file = 'output.xlsx'
    try:
        # Excelファイルの存在確認とWorkbookオブジェクトの生成
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # デフォルトシートの削除

        # ホスト名のシートが存在するか確認し、必要に応じて作成
        if hostname in wb.sheetnames:
            ws = wb[hostname]
        else:
            ws = wb.create_sheet(title=hostname)
            headers = ["Timestamp", "Interface", "input packets", "output packets", "input unicasts", "output unicasts", "input broadcast packets", "output broadcast packets", "input multicast packets", "output multicast packets"]
            ws.append(headers)

        # データの書き込み
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        for info in interface_info:
            ws.append([timestamp] + list(info))

        #フィルターをONにする
        ws.auto_filter.ref = f"A1:J{len(interface_info) + 1}"

        # 変更を保存
        wb.save(excel_file)
    except PermissionError as e:
        logging.error(f"エラー: ファイル '{excel_file}' への書き込み権限がありません。詳細: {e}")
        print(f"エラー: ファイル '{excel_file}' への書き込み権限がありません。詳細: {e}")
        sys.exit(1)
    except InvalidFileException as e:
        logging.error(f"エラー: ファイル '{excel_file}' は無効または破損しています。詳細: {e}")
        print(f"エラー: ファイル '{excel_file}' は無効または破損しています。詳細: {e}")
        sys.exit(1)
    except Exception as e:
        logging.error(f"エラー: {e}")
        print(f"エラー: {e}")
        sys.exit(1)

# メイン関数
def main():
    username, password = get_credentials()  # 認証情報の取得
    jump_host_ip = '10.10.0.222'
    jump_host_user, jump_host_pass = username, password  # 踏み台ホストの認証情報

    # ホスト情報の読み込み
    hosts = load_hosts()

    for hostname, ip_address in hosts:
        json_output = ssh_via_jump_host(jump_host_ip, jump_host_user, jump_host_pass, ip_address,
                                        username, password)
        if json_output:  # JSON出力がある場合のみ処理を続行
            interface_info = extract_interface_info(json.loads(json_output))
            write_to_excel(hostname or ip_address, interface_info)
            print(f"ホスト {hostname or ip_address} のデータをエクセルに書き込みました。")
if __name__ == "__main__":
    main()
